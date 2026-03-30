"""
generate_sample_excel.py
========================
Run this script once to create a sample Excel file for testing the
intelligent import endpoint.

Usage (from the backend directory):
    python generate_sample_excel.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Students"

HEADERS = [
    "Name", "Email", "Department", "Year", "Section",
    "RollNumber", "Semester", "Subject", "Marks", "Attendance",
    "Backlogs", "Detained",
]

# Style headers
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2563EB")
header_align = Alignment(horizontal="center", vertical="center")

for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

# Sample data rows  (year 1-4, semester 1-8)
ROWS = [
    # (Name, Email, Dept, Year, Section, Roll, Sem, Subject, Marks, Attendance, Backlogs, Detained)
    ("Alice Johnson",  "alice@college.edu",  "CSE", 1, "A", "UP1AABB01", 1, "Mathematics I",           88.5, 91.0, 0, False),
    ("Alice Johnson",  "alice@college.edu",  "CSE", 1, "A", "UP1AABB01", 1, "Programming Fundamentals", 76.0, 85.0, 0, False),
    ("Bob Smith",      "bob@college.edu",    "ECE", 2, "B", "UP2CCDD02", 3, "Digital Circuits",         65.0, 72.5, 1, False),
    ("Carol Davis",    "carol@college.edu",  "CSE", 3, "A", "UP3EEFF03", 5, "Operating Systems",        55.0, 60.0, 2, True),
    ("David Lee",      "david@college.edu",  "ME",  4, "C", "UP4GGHH04", 7, "Thermodynamics",           90.0, 95.0, 0, False),
    ("Eva Martin",     "eva@college.edu",    "CSE", 2, "A", "UP2IIJJ05", 3, "Data Structures",          72.0, 78.0, 0, False),
    ("Frank Wilson",   "frank@college.edu",  "ECE", 1, "B", "UP1KKLL06", 2, "Circuit Theory",           80.0, 88.0, 0, False),
    ("Grace Kim",      "grace@college.edu",  "CSE", 4, "A", "UP4MMNN07", 7, "Machine Learning",         95.0, 97.0, 0, False),
    # Duplicate email on purpose — should be processed correctly (idempotency demo)
    ("Alice Johnson",  "alice@college.edu",  "CSE", 1, "A", "UP1AABB01", 1, "Mathematics I",           88.5, 91.0, 0, False),
]

for row_data in ROWS:
    ws.append(row_data)

# Column widths
col_widths = [18, 28, 8, 6, 8, 14, 8, 26, 8, 12, 10, 10]
for col_idx, w in enumerate(col_widths, start=1):
    ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = w

wb.save("sample_import.xlsx")
print("✅ sample_import.xlsx created — upload it to POST /api/v1/import/excel")
